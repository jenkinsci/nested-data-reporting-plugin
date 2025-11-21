import openpyxl

workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.title = "Sheet1"
sheet1.cell(row=1, column=1, value="ID")
sheet1.cell(row=1, column=2, value="Value")
sheet1.cell(row=2, column=1, value="1")
sheet1.cell(row=2, column=2, value="10")
sheet1.cell(row=3, column=1, value="2")
sheet1.cell(row=3, column=2, value="20")

sheet2 = workbook.create_sheet("Sheet2")
sheet2.cell(row=1, column=1, value="ID")
sheet2.cell(row=1, column=2, value="Value")
sheet2.cell(row=2, column=1, value="3")
sheet2.cell(row=2, column=2, value="30")
sheet2.cell(row=3, column=1, value="4")
sheet2.cell(row=3, column=2, value="40")

workbook.save("src/test/resources/test_multi.xlsx")

import openpyxl

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Sheet1"
sheet.cell(row=1, column=1, value="ID")
sheet.cell(row=1, column=2, value="Value")
sheet.cell(row=2, column=1, value="1")
sheet.cell(row=2, column=2, value="10")
sheet.cell(row=3, column=1, value="2")
sheet.cell(row=3, column=2, value="20")
workbook.save("src/test/resources/test.xlsx")
